import openpyxl
import random  # Import the random module

# Define the values for the second column in the exact order
second_column_values = [
    "Boolean",
    "Date",
    "DateTime",
    "Integer",
    "Number",
    "Text",
    "Integer",
    "List",
    "URL",
    "Percent"
]
# Define the values for the third column (names)
names = ['chamal', 'joshua', 'ethan', 'kiara']

# Define the number of rows you want
num_rows = 200  # Change this value to the desired number of rows

# Create a new workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Generate the content for the specified number of rows
for i in range(1, num_rows + 1):
    # First column: Field_1, Field_2, ...
    first_column_value = f"Field_{i}"

    # Second column: Values in exact order
    second_column_value = second_column_values[(i - 1) % len(second_column_values)]

    # Third column: Random name
    third_column_value = random.choice(names)

    # Write values to the worksheet
    sheet.cell(row=i, column=1, value=first_column_value)
    sheet.cell(row=i, column=2, value=second_column_value)
    sheet.cell(row=i, column=3, value=third_column_value)

# Save the workbook to a file
workbook.save("dataset_200_fieldtypes.xlsx")


print('Finished!')